"""CMS NCCI PTP edits — pulls the quarterly practitioner edit ZIP.

We:
  1. Pull the landing HTML and find the most recent practitioner-PTP zip link.
  2. Download the zip and extract every PTP edit it contains:
       - the *tab-delimited* additions/deletions files (NCCI calls them .txt but
         they're TSV, not CSV), and
       - any .xlsx workbooks, parsed with a streaming reader.
  3. Filter rows to HCPCS codes Engine 1 evaluates (OTP G/H-codes).

Files inside the delta ZIP look like:
    MCR_NCCI_Additions_Eff_<QTR>.txt
    MCR_NCCI_Deletions_Eff_<QTR>.txt
    MCR_NCCI_CCMIChgs_Eff_<QTR>.txt
    MCR_NCCI_Changes_Eff_<QTR>.xlsx

xlsx handling (review action: "ingest the full PTP universe"): the delta zip only
carries that quarter's additions/deletions, so Engine 1 would miss historical
unbundling edits (e.g. the H0020 + 80305 conflict). The *full* quarterly PTP
tables CMS publishes are xlsx workbooks of ~2.5M rows. We read those with
openpyxl in read_only mode and stream row-by-row via iter_rows() so memory stays
flat regardless of table size — never loading the whole sheet into RAM. The same
code path also reads the delta zip's Changes.xlsx; rows are de-duplicated against
the .txt additions/deletions so we don't double-count.
"""
from __future__ import annotations

import csv
import hashlib
import io
import re
import zipfile
from urllib.parse import parse_qs, urljoin, urlparse

import openpyxl
from bs4 import BeautifulSoup

from .base import BaseScraper, ScrapeResult

# HCPCS codes Engine 1 evaluates. Expanded to cover the full OTP weekly bundle
# range (G2067–G2080) plus take-home naloxone codes (G1028, G2215) and the
# core FL Medicaid H-codes.
RELEVANT_CODES = {
    # OTP weekly bundle G-codes (CMS Pub 100-04 Ch 39)
    "G2067", "G2068", "G2069", "G2070", "G2071", "G2072", "G2073",
    "G2074", "G2075", "G2076", "G2077", "G2078", "G2079", "G2080",
    # OTP intake / add-on / take-home
    "G2086", "G2087", "G2088",
    "G1028", "G2215",          # take-home naloxone
    "G0137",                   # IOP threshold trigger
    # FL Medicaid MAT H-codes
    "H0001", "H0004", "H0005", "H0006", "H0020", "H0033", "H0047",
    "H0050", "H2010", "H2017",
}

# Full practitioner PTP table parts. CMS splits the ~2.6M-row practitioner
# universe into f1..fN zips because each exceeds Excel's 1,048,576-row cap, and
# fronts each download with an AMA-license click-through:
#   /license/ama?file=/files/zip/medicare-ncci-2026q2-practitioner-ptp-edits-ccipra-v321r0-f1.zip
# "ccipra" = Correct Coding Initiative, PRActitioner ("ccioph" is the hospital
# set, which we don't evaluate). We match the practitioner full-table parts and
# pull every part of the most recent quarter+version.
RE_FULL_PRACTITIONER = re.compile(
    r"medicare-ncci-(?P<year>\d{4})q(?P<qtr>\d)-practitioner-ptp-edits-"
    r"ccipra-v(?P<vmaj>\d+)r(?P<vmin>\d+)-f(?P<part>\d+)\.zip",
    re.I,
)
# The quarterly delta (additions/deletions/revisions). Kept as a fallback when
# the full table can't be located, so a run still produces *something*.
RE_DELTA_PRACTITIONER = re.compile(
    r"practitioner.*(?:additions|deletions|revisions|quarterly).*\.zip", re.I
)


class NCCIScraper(BaseScraper):
    def parse(self) -> ScrapeResult:
        html = self.fetch_text(self.source.url)
        soup = BeautifulSoup(html, "lxml")
        hrefs = [a["href"] for a in soup.find_all("a", href=True)]

        full_parts = _select_full_table_parts(hrefs, self.source.url)
        if full_parts:
            return self._scrape_full_table(full_parts)

        # Fallback: the quarterly delta zip (current-quarter changes only).
        delta = next(
            (urljoin(self.source.url, h) for h in hrefs if RE_DELTA_PRACTITIONER.search(h)),
            None,
        )
        if delta is None:
            return ScrapeResult(
                source_key=self.source.key,
                source_name=self.source.name,
                fetched_at="",
                content_sha256="",
                raw_path="",
                parsed={"doc": "NCCI", "edits": []},
                warnings=["No practitioner PTP zip link (full table or delta) found"],
            )
        payload = self.fetch_bytes(delta)
        raw_path, digest = self._persist_raw(payload, suffix=".zip")
        edits = self._extract_relevant_edits(payload)
        return ScrapeResult(
            source_key=self.source.key,
            source_name=self.source.name,
            fetched_at="",
            content_sha256=digest,
            raw_path=str(raw_path),
            parsed={
                "doc": "NCCI PTP Edits (practitioner, DELTA fallback)",
                "table_scope": "delta",
                "source_zips": [delta],
                "edit_count_relevant": len(edits),
                "edits": edits,
            },
            warnings=["Full PTP table not found; fell back to quarterly delta zip"],
        )

    def _scrape_full_table(self, parts: list[str]) -> ScrapeResult:
        """Download + stream-parse every part of the full practitioner table."""
        edits: list[dict] = []
        part_digests: list[str] = []
        first_raw_path = ""
        warnings: list[str] = []
        seen: set[tuple[str, str, str]] = set()
        for url in parts:
            payload = self.fetch_bytes(url)
            raw_path, digest = self._persist_raw(payload, suffix=".zip")
            part_digests.append(digest)
            first_raw_path = first_raw_path or str(raw_path)
            # Guard against the license-page HTML coming back instead of a zip.
            if not zipfile.is_zipfile(io.BytesIO(payload)):
                warnings.append(f"Not a zip (license redirect?): {url}")
                continue
            self._extract_relevant_edits(payload, edits=edits, seen=seen)
        # Stable combined digest so the diff-checker tracks the table as a unit.
        combined = hashlib.sha256("".join(part_digests).encode()).hexdigest()
        return ScrapeResult(
            source_key=self.source.key,
            source_name=self.source.name,
            fetched_at="",
            content_sha256=combined,
            raw_path=first_raw_path,
            parsed={
                "doc": "NCCI PTP Edits (practitioner, FULL table)",
                "table_scope": "full",
                "source_zips": parts,
                "part_count": len(parts),
                "edit_count_relevant": len(edits),
                "edits": edits,
            },
            warnings=warnings,
        )

    @classmethod
    def _extract_relevant_edits(
        cls,
        payload: bytes,
        edits: list[dict] | None = None,
        seen: set[tuple[str, str, str]] | None = None,
    ) -> list[dict]:
        # De-dup across files: the delta zip ships the same edits as both .txt and
        # the Changes.xlsx, and the full table's split parts can repeat a pair.
        # Callers parsing multiple zips (the full table) pass a shared edits/seen
        # so de-duplication spans every part.
        if edits is None:
            edits = []
        if seen is None:
            seen = set()
        with zipfile.ZipFile(io.BytesIO(payload)) as zf:
            for name in zf.namelist():
                low = name.lower()
                if low.endswith(".txt") and ("additions" in low or "deletions" in low):
                    kind = "addition" if "additions" in low else "deletion"
                    with zf.open(name) as fh:
                        text = io.TextIOWrapper(fh, encoding="latin-1", errors="ignore")
                        # NCCI .txt files are TAB-delimited, not comma. First row is
                        # a copyright blurb, then a header row, then data rows:
                        # <Column1>\t<Column2>\t<ModifierIndicator>\t<...>.
                        cls._collect(
                            csv.reader(text, delimiter="\t"), name, kind, edits, seen
                        )
                elif low.endswith(".xlsx"):
                    cls._collect_xlsx(zf.read(name), name, edits, seen)
        return edits

    @staticmethod
    def _collect(rows, source_file, default_kind, edits, seen) -> None:
        """Filter an iterable of rows to relevant OTP edits, de-duplicating."""
        for row in rows:
            if row is None or len(row) < 2:
                continue
            c1 = "" if row[0] is None else str(row[0]).strip()
            c2 = "" if row[1] is None else str(row[1]).strip()
            # Skip header / copyright rows (don't look like HCPCS codes).
            if not _looks_like_hcpcs(c1) or not _looks_like_hcpcs(c2):
                continue
            if c1 not in RELEVANT_CODES and c2 not in RELEVANT_CODES:
                continue
            key = (c1, c2, default_kind)
            if key in seen:
                continue
            seen.add(key)
            # The modifier indicator (0/1/9) sits at a different column depending on
            # layout: index 2 in the delta files, index 5 in the full PTP table
            # (which also has *, effective/deletion dates and a rationale column).
            # Scan past the code pair for the cell that is exactly 0/1/9.
            modifier = ""
            for cell in row[2:]:
                if cell is None:
                    continue
                token = str(cell).strip().split("\n")[0].strip()
                if token in ("0", "1", "9"):
                    modifier = token
                    break
            edits.append(
                {
                    "column1": c1,
                    "column2": c2,
                    "modifier_indicator": modifier,
                    "edit_kind": default_kind,
                    "source_file": source_file,
                }
            )

    @classmethod
    def _collect_xlsx(cls, data: bytes, member_name: str, edits, seen) -> None:
        """Stream an xlsx workbook row-by-row (read_only) into the edit list.

        read_only=True + iter_rows() keeps memory flat for the multi-million-row
        full PTP tables; openpyxl never materializes the whole sheet.
        """
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                kind = _xlsx_sheet_kind(ws.title, member_name)
                cls._collect(
                    ws.iter_rows(values_only=True),
                    f"{member_name}#{ws.title}",
                    kind,
                    edits,
                    seen,
                )
        finally:
            wb.close()


# Medicaid NCCI PTP zips. CMS posts the Medicaid edit files on a separate page
# with a different naming convention from the Medicare practitioner table — the
# H0020 + 80305 (and other H-code) unbundling edits live here, not in the
# Medicare file. We match practitioner PTP zips and prefer them over the
# outpatient-hospital set (which Engine 1 doesn't evaluate). Naming has drifted
# across quarters, so the match is deliberately loose: any .zip whose link text
# mentions Medicaid + PTP. MUE (medically-unlikely-edit) files are excluded.
RE_MEDICAID_PTP = re.compile(
    r"(?=.*medicaid)(?=.*ptp)(?!.*\bmue\b)(?!.*outpatient).*\.zip", re.I
)


class MedicaidNCCIScraper(NCCIScraper):
    """Medicaid NCCI PTP edits — the H-code (H0020) unbundling source.

    Reuses NCCIScraper's zip/xlsx extraction and OTP-code filtering wholesale;
    only the link discovery differs (Medicaid edit-files page vs. the Medicare
    practitioner table). Degrades gracefully — if no Medicaid PTP zip link is
    found on the page, it returns an empty edit set with a warning rather than
    raising, so a layout change on the CMS side never takes down the run.
    """

    def parse(self) -> ScrapeResult:
        html = self.fetch_text(self.source.url)
        soup = BeautifulSoup(html, "lxml")
        hrefs = [a["href"] for a in soup.find_all("a", href=True)]

        # Collect every Medicaid PTP zip, resolve any AMA-license wrapper, dedupe.
        zip_urls: list[str] = []
        seen_urls: set[str] = set()
        for h in hrefs:
            if not RE_MEDICAID_PTP.search(h):
                continue
            url = _resolve_cms_href(h, self.source.url)
            if url not in seen_urls:
                seen_urls.add(url)
                zip_urls.append(url)

        if not zip_urls:
            return ScrapeResult(
                source_key=self.source.key,
                source_name=self.source.name,
                fetched_at="",
                content_sha256="",
                raw_path="",
                parsed={"doc": "Medicaid NCCI PTP Edits", "edits": []},
                warnings=["No Medicaid PTP zip link found on the edit-files page"],
            )

        edits: list[dict] = []
        seen: set[tuple[str, str, str]] = set()
        part_digests: list[str] = []
        first_raw_path = ""
        warnings: list[str] = []
        for url in zip_urls:
            payload = self.fetch_bytes(url)
            raw_path, digest = self._persist_raw(payload, suffix=".zip")
            part_digests.append(digest)
            first_raw_path = first_raw_path or str(raw_path)
            if not zipfile.is_zipfile(io.BytesIO(payload)):
                warnings.append(f"Not a zip (license redirect?): {url}")
                continue
            self._extract_relevant_edits(payload, edits=edits, seen=seen)

        combined = hashlib.sha256("".join(part_digests).encode()).hexdigest()
        return ScrapeResult(
            source_key=self.source.key,
            source_name=self.source.name,
            fetched_at="",
            content_sha256=combined,
            raw_path=first_raw_path,
            parsed={
                "doc": "Medicaid NCCI PTP Edits (practitioner)",
                "table_scope": "medicaid_ptp",
                "source_zips": zip_urls,
                "edit_count_relevant": len(edits),
                "edits": edits,
            },
            warnings=warnings,
        )


def _resolve_cms_href(href: str, base_url: str) -> str:
    """Turn an AMA-license-wrapped href into the direct file URL.

    `/license/ama?file=/files/zip//medicare-...zip` -> the `file=` target, with
    the occasional doubled slash normalized, joined onto the CMS origin.
    """
    parsed = urlparse(href)
    if parsed.path.endswith("/license/ama"):
        target = parse_qs(parsed.query).get("file", [""])[0]
        if target:
            href = re.sub(r"/{2,}", "/", target)
    return urljoin(base_url, href)


def _select_full_table_parts(hrefs: list[str], base_url: str) -> list[str]:
    """Pick all parts of the most recent full practitioner PTP table.

    Returns resolved (license-unwrapped) URLs ordered by part number, or [] when
    no full-table links are present (caller then falls back to the delta zip).
    """
    matches = []
    for h in hrefs:
        m = RE_FULL_PRACTITIONER.search(h)
        if m:
            matches.append((m, h))
    if not matches:
        return []
    # Newest quarter, then newest version. The parts of one release share these.
    def release_key(item):
        m = item[0]
        return (int(m["year"]), int(m["qtr"]), int(m["vmaj"]), int(m["vmin"]))

    best = max(release_key(item) for item in matches)
    chosen = [item for item in matches if release_key(item) == best]
    chosen.sort(key=lambda item: int(item[0]["part"]))
    # De-dupe parts (a page can list the same href twice) while preserving order.
    seen_parts: set[int] = set()
    urls: list[str] = []
    for m, h in chosen:
        part = int(m["part"])
        if part in seen_parts:
            continue
        seen_parts.add(part)
        urls.append(_resolve_cms_href(h, base_url))
    return urls


_HCPCS_RE = re.compile(r"^[A-Z0-9]{5}$")


def _looks_like_hcpcs(code: str) -> bool:
    """HCPCS codes are 5 chars, alphanumeric (e.g., G2067, 99213, 0395T)."""
    return bool(_HCPCS_RE.match(code))


def _xlsx_sheet_kind(sheet_title: str, member_name: str) -> str:
    """Classify an xlsx sheet by its name.

    Delta workbooks name sheets NCCI_Adds_*/NCCI_Dels_*/NCCI_CCMIChgs_*. Full PTP
    tables aren't additions/deletions at all — they're the standing edit universe,
    so anything we can't classify falls back to "ptp_edit".
    """
    name = f"{sheet_title} {member_name}".lower()
    if "add" in name:
        return "addition"
    if "del" in name:
        return "deletion"
    if "ccmi" in name or "chg" in name or "change" in name:
        return "ccmi_change"
    return "ptp_edit"
